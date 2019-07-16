# Bulk SMS Reminder System with loading data files from Excel Files
'''
# IMPORTANT:
1) User must install Airmore Apps on their Android or IOS phone
2) After run the program, connect and accept connection from phone.
3) Once connected to phone, load the excel file into the program...(only .XLSX file format will be accepted)
4) The software will send the sms if at the column of "To Send" is mark as "y" or "Y", and also the expired date is within the days specified in config.ini file.
5) After successfully send sms, the "SMS" will mark as "OK", otherwise will be blank.
6) In most country, sending sms will be controlled under fair use policy, hence it is recommended not to send over 100 sms per day.
7) This program will be best used if you have signed unlimited sms package with your local telco.
8) Current Market Price for send bulk sms will be at USD 0.02 per sms.
9) This source code and program is used at your own risk, author do not will not take any responsible for your lost during this program usage.
10) No guarantee or warranty will be provided.

# Author: Edmond Lam
# Completed date: 16-Jul-2019
'''
import tkinter as tk
from tkinter import PhotoImage
from tkinter import ttk
from tkinter import filedialog
from tkinter import StringVar
import webbrowser

import re
import openpyxl
from openpyxl import load_workbook
import os.path
import configparser
import time
from datetime import datetime
from datetime import date
import sys

from time import gmtime,strftime

from ipaddress import IPv4Address  # for your IP address
from pyairmore.request import AirmoreSession  # to create an AirmoreSession
from pyairmore.services.messaging import MessagingService  # to send messages


#Appearance Setting=======================================================

LARGE_FONT="Helvetica 22 bold"
BTN_FONT=("Helvetica", 20, 'bold')
TREE_FONT=("Helvetica", 18, 'bold')
KEY_FONT="Helvetica 35 bold"
UID_FONT=('Consolas', 38, 'bold')
StandardWinSize="850x530+50+50"
strURLstrURL = "web.airmore.com"

config = configparser.ConfigParser()
config.read('config.ini')
delaysec=config['settings']['delayseconds']
host = config['settings']['url']
expday = config['settings']['expire']
oneurl=str(host)


class smsapp(tk.Tk):
	
	def __init__(self,*args, **kwargs):
	
		tk.Tk.__init__(self, *args, **kwargs)

		tk.Tk.wm_title(self,"Win Fon Trading SMS Reminder System 1.1")

		self.shared_icon ={
				"Send"	:	tk.PhotoImage(file="./image/icon/32x32/32send.png"),	
				"Excel"	:	tk.PhotoImage(file="./image/icon/32x32/excel.png"),
				"SMS"	:	tk.PhotoImage(file="./image/icon/32x32/QR.png"),
				"Phone"	:	tk.PhotoImage(file="./image/icon/32x32/phone.png"),
			} 
			
		container = tk.Frame(self)
		container.grid(row=0, column=0)
		
		self.frames ={}
		
		for F in (StartPage,NetworkForm):#
			
			frame = F(parent=container,controller=self)
			self.frames[F] = frame
			frame.config(width=850, height=530)#bg="black"
			frame.grid(row=0, column =0, sticky="nsew")
		
		self.show_frame(StartPage)
		
	def show_frame(self, contnr):
		
		frame=self.frames[contnr]
		frame.tkraise()

		
	def get_frame(self,contnr):
		frame=self.frames[contnr]
		return frame

class StartPage(tk.Frame):
	global session, service, workbook,worksheet, row_count,host,wb,host
	datelabel=''
	timelabel=''
	saatlabel=''
	uidlabel=''
	
	def __init__(self, parent,controller):
		tk.Frame.__init__(self,parent)
		self.controller=controller

		self.textEntryVar=StringVar()
		self.LabelTextVar=StringVar()
		self.gpw_pb_ivar = tk.IntVar()

		self.labeldate =tk.Label(self, text=self.datelabel, font=('Consolas', 23, 'bold'),height="1", width="30", justify=tk.LEFT,fg='white',bg='black')
		self.labeldate.grid(row=0,column=0,pady=1,padx=10, sticky="W")
		
		self.gap = tk.Label(self,text="Phone IP: ",font="Helvetica 16 bold",height="1")#,bg='black')
		self.gap.grid(row=1,column=0, sticky='W', columnspan=1, ipady=1)
		
		self.e = tk.Entry(self,textvariable=self.textEntryVar, font="Helvetica 22 bold", width=15, justify="center")#for the figure entered and to use globally
		self.e.grid(row=2, column=0, sticky='W',columnspan=3, ipady=5)
		self.e.focus()
		
		self.btnconnect = tk.Button(self, command=lambda:self.launchurl(), width =200,text="Connect Phone", font="Helvetica 16 bold",compound=tk.TOP,image=self.controller.shared_icon["Phone"])#, height=50,width=150)
		self.btnconnect.grid(row=2, column=1, sticky="W")

		
		#Just a gap========================================================
		self.gap = tk.Label(self,height="1")
		self.gap.grid(row=2,column=0, columnspan=1, ipady=1)

		self.gap1 = tk.Label(self,text="Status: ",font="Helvetica 16 bold")
		self.gap1.grid(row=3,column=0,sticky="W")
		
		self.nstatus = tk.Label(self, text="Please open your phone Airmore Apps",font="Helvetica 16 bold", anchor='w')#,bg="black", fg='white')
		self.nstatus.grid(row=4, column=0,sticky="W")

		
		self.gap = tk.Label(self,height="1")
		self.gap.grid(row=2,column=0, columnspan=1, ipady=1)
		
		self.btnexcel = tk.Button(self, command=lambda:self.loadexcel(), width =350,text="Load Excel Files and Send SMS", font="Helvetica 16 bold",compound=tk.TOP,image=self.controller.shared_icon["Excel"])#, height=50,width=150)
		self.btnexcel.grid(row=5, column=0, sticky="W")
		
		self.fname = tk.Label(self, text="Filename :",font="Helvetica 16 bold", anchor='w')#,bg="black", fg='white')
		self.fname.grid(row=6, column=0,sticky="W")
		
		self.sheetname = tk.Label(self, text="Worksheet :",font="Helvetica 16 bold", anchor='w',fg='Blue')#,bg="black", fg='white')
		self.sheetname.grid(row=7, column=0,sticky="W")
		
		self.record = tk.Label(self, text="Total SMS ready to send: 0",font="Helvetica 16 bold", anchor='w')#,bg="black", fg='white')
		self.record.grid(row=8, column=0,sticky="W")

		self.smscounter = tk.Label(self, text="Total SMS Successfully sent: 0",font="Helvetica 16 bold", anchor='w',fg='Green')#,bg="black", fg='white')
		self.smscounter.grid(row=9, column=0,sticky="W")
		
		self.failcounter = tk.Label(self, text="Total SMS Fail to send: 0",font="Helvetica 16 bold", anchor='w',fg='red')#,bg="black", fg='white')
		self.failcounter.grid(row=10, column=0,sticky="W")

		self.btnsend = tk.Button(self, command=lambda:webbrowser.open(strURLstrURL), width =200,text="QR Code Login", font="Helvetica 16 bold",compound=tk.TOP,image=self.controller.shared_icon["SMS"])#, height=50,width=150)
		self.btnsend.grid(row=1, column=1, sticky="W")

		
		self.copyright1 = tk.Label(self, text="(C) Twinwood Smartech 2019            ", font=('Consolas', 10, 'bold'))#,bg='black',fg='white')#fg="red"
		self.copyright1.grid(row=20,column=0, sticky="W")
		
		self.debugstatus = tk.Label(self, text='',font="Helvetica 16 bold")#fg="red"
		self.debugstatus.grid(row=21,column=0, columnspan=3, ipady=5,sticky='W')
		
		self.progressBar = ttk.Progressbar(self,
									mode='determinate',
									maximum=650,
									length=650,
									variable=self.gpw_pb_ivar)
		self.progressBar.grid(row=13,column=0, columnspan=3, ipady=5)
		
		self.textEntryVar.set(str(host))
		self.getTime()
		

			
	def loadexcel(self):
		global workbook,worksheet,row_count,wb
		try:

			self.filename = filedialog.askopenfilename()
			#self.df = pd.read_excel(self.filename)

			self.fname['text']="Filename: "+str(self.filename)
			self.fname['fg']="Blue"
			self.record['text']="Total SMS ready to send: 0"
			self.smscounter['text']="Total SMS Successfully sent:0"
			self.smscounter['fg']="Green"
			self.failcounter['text']="Total SMS Fail to send: 0"
			self.failcounter['fg']="Red"
			self.nstatus['text']="Processing..."
			
			workbook=openpyxl.load_workbook(self.filename)
			#print(workbook.sheetnames)
			worksheet = workbook.active
			#print(worksheet)
			self.sheetname['text']=str(worksheet)
			row_count = worksheet.max_row
				
			self.record['text']="Total record found: "+ str(row_count-1)


			#except Exception as e:
			#	self.nstatus['text']="Excel File Error!"
			#	self.nstatus['fg']="red"
			#	print(e)
			messageRT_BC="Hi {name}, your vehicle number {vehicle} registration is going to expire soon. Please contact Michael via this number. TQ"

			#print (self.df)
			#Name column(i,0)
			#Telephone No. column(i,1)
			#Exdate column(i,2)
			#Vehicle No. column(i,3)
			self.nstatus['text']="Sending All SMS...."
			self.nstatus['fg']="Red"
			sms_counter=0
			fail_counter=0
			wantsend_counter=0
			date_format="%Y-%m-%d %H:%M:%S"
			for i in range(2,int(row_count)+1,1):
				SMS=worksheet.cell(row=i, column=8).value
				ToSend=	worksheet.cell(row=i, column=7).value
				if (ToSend is "Y" or ToSend is "y") and SMS is None:
					wantsend_counter+=1
			self.record['text']="Total SMS Ready to send: "+ str(wantsend_counter)

			for i in range(2,int(row_count)+1,1):

				self.gpw_pb_ivar.set(int((i/row_count)*650))
				self.progressBar.update() # this fixes the problem
				try:
					name= worksheet.cell(row=i, column=1).value
					number= worksheet.cell(row=i, column=2).value
					rawdate= worksheet.cell(row=i, column=3).value
					vehicle= worksheet.cell(row=i, column=4).value
					Language=worksheet.cell(row=i, column=5).value
					Item= worksheet.cell(row=i, column=6).value
					ToSend=	worksheet.cell(row=i, column=7).value
					SMS=worksheet.cell(row=i, column=8).value

					if rawdate is not None:
						exdate1=datetime.strptime(str(rawdate),date_format)
						diff=exdate1-datetime.now()
						exdate=str(exdate1).replace(" 00:00:00", "")
					else:
						exdate=None
					
					if (ToSend is "Y" or ToSend is "y") and SMS is None and int(diff.days)<int(expday)and int(diff.days)>0:
						if str(Item)=="RT":
							if number is not None and name is not None and vehicle is not None and rawdate is not None:
								if str(Language)=="BC":
									smessage=messageRT_BC.format(name=str(name),vehicle=str(vehicle),exdate=str(exdate))
									service.send_message(number, smessage)
									worksheet.cell(i,8).value="OK"
									sms_counter+=1

						elif str(Item)=="GDL":
							if number is not None and name is not None and rawdate is not None:
								if str(Language)=="BC":
									smessage=messageGDL_BC.format(name=str(name),exdate=str(exdate))
									service.send_message(number, smessage)
									worksheet.cell(i,8).value="OK"
									sms_counter+=1

						time.sleep(int(delaysec))

					else:
						fail_counter+=1
			
				except Exception as e:
					fail_counter+=1
					self.nstatus['text']="Send SMS error!"
					self.nstatus['fg']="red"
					print(e)

			workbook.save(self.filename)
			self.smscounter['text']="Total SMS Successfully Sent:"+str(sms_counter)
			self.smscounter['fg']="Green"
			self.failcounter['text']="Total SMS Failed to Send:"+str(wantsend_counter-sms_counter)#fail_counter)
			self.failcounter['fg']="Red"
			self.nstatus['text']="All SMS Delivered!"
			self.nstatus['fg']="Green"	
		except Exception as e:
			self.nstatus['text']="Try load file again!"
			self.nstatus['fg']="blue"

	
	def launchurl(self):
		global session, service,oneurl

		with open("config.ini", "r+") as f:
			f.seek(0)
			d = f.readlines()
			#print(d)
			for i in d:
				if "url="+oneurl+'\n' not in i:
					f.write(i)
			f.write("url="+self.textEntryVar.get()+"\n")
			f.truncate()

		try:
			url= str(self.textEntryVar.get())
			ip = IPv4Address(url)# let's create an IP address object
			# now create a session
			session = AirmoreSession(ip)
			# if your port is not 2333
			# session = AirmoreSession(ip, 2334)  # assuming it is 2334

			was_accepted = session.request_authorization()
			
			if was_accepted==True:
				self.nstatus['text']="Connected to Phone!"
				self.nstatus['fg']="blue"
				service = MessagingService(session)


		except Exception as e:
			self.nstatus['text']="IP Address Error!"
			self.nstatus['fg']="red"

	def getTime(self):
		global host
		self.datelabel = strftime(" %H:%M  %A, %d %B %Y ")
		self.labeldate.configure(text=self.datelabel)
		self.after(9000, self.getTime)               #wait 0.5 sec and go again	
	
	
	def progressBar(self):
		
		self.progress_bar=ttk.Progressbar(self, orient="horizontal",length =650, mode ="determinate")
		self.progress_bar.grid(row=13,column=0, columnspan=3, ipady=5)


	def start_progress(self):
		self.progress_bar["value"]+=20
		self.after(500,self.start_progress)	
        
class NetworkForm(tk.Frame): #For others functions
	def __init__(self, parent,controller):
		tk.Frame.__init__(self,parent)
		self.controller=controller
		
		self.toolgap = tk.Label(self,text="",bg='black', font="Helvetica 16 bold")
		self.toolgap.grid(row=1,column=0, columnspan=3, ipady=5,sticky="W")
		
		self.gapTitle = tk.Label(self,text="Network" ,height="1", width="35", fg="white", bg="blue",font="Helvetica 28 bold")
		self.gapTitle.grid(row=2,column=0, columnspan=50, ipady=5,sticky='W')


if __name__=='__main__':

	app = smsapp()
	app.geometry(StandardWinSize)

	app.mainloop()
